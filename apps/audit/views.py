import logging
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import Http404
from django.views.generic import ListView, View

from .forms import AuditLogFilterForm
from .models import AuditLog

logger = logging.getLogger(__name__)


class AdminRequiredMixin:
    """Только admin — иначе 404."""
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.role != "admin":
            raise Http404
        return super().dispatch(request, *args, **kwargs)


class AuditLogListView(AdminRequiredMixin, LoginRequiredMixin, ListView):
    template_name = "audit/list.html"
    context_object_name = "logs"
    paginate_by = 50

    def get_queryset(self):
        qs = AuditLog.objects.select_related("user").order_by("-created_at")
        form = AuditLogFilterForm(self.request.GET)
        if not form.is_valid():
            return qs

        cd = form.cleaned_data
        if cd.get("user_search"):
            q = cd["user_search"]
            from django.db.models import Q
            qs = qs.filter(
                Q(user__username__icontains=q)
                | Q(user__first_name__icontains=q)
                | Q(user__last_name__icontains=q)
            )
        if cd.get("action"):
            qs = qs.filter(action__icontains=cd["action"])
        if cd.get("entity_type"):
            qs = qs.filter(entity_type__icontains=cd["entity_type"])
        if cd.get("date_from"):
            qs = qs.filter(created_at__date__gte=cd["date_from"])
        if cd.get("date_to"):
            qs = qs.filter(created_at__date__lte=cd["date_to"])
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter_form"] = AuditLogFilterForm(self.request.GET or None)
        return context


class AuditLogExportView(AdminRequiredMixin, LoginRequiredMixin, View):
    """GET /audit/export/?format=xlsx — экспорт лога в XLSX."""

    def get(self, request):
        from apps.reports.exporters import export_xlsx
        from .forms import AuditLogFilterForm

        form = AuditLogFilterForm(request.GET)
        filters = {}
        if form.is_valid():
            cd = form.cleaned_data
            if cd.get("date_from"):
                filters["date_from"] = cd["date_from"]
            if cd.get("date_to"):
                filters["date_to"] = cd["date_to"]

        # Получаем данные напрямую, не через report services
        qs = AuditLog.objects.select_related("user").order_by("-created_at")
        if filters.get("date_from"):
            qs = qs.filter(created_at__date__gte=filters["date_from"])
        if filters.get("date_to"):
            qs = qs.filter(created_at__date__lte=filters["date_to"])

        return _export_audit_xlsx(qs, request.user)


def _export_audit_xlsx(qs, user):
    """Экспортирует AuditLog QS в XLSX-ответ."""
    import io
    from datetime import date
    from django.http import HttpResponse
    from apps.audit.services import audit_log as write_audit

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse("openpyxl не установлен", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лог системы"

    headers = ["Дата и время", "Пользователь", "Действие", "Тип объекта", "ID объекта", "IP", "Детали"]
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    for log in qs[:10000]:
        ws.append([
            log.created_at.strftime("%d.%m.%Y %H:%M:%S") if log.created_at else "",
            log.user.username if log.user else "—",
            log.action,
            log.entity_type,
            log.entity_id or "",
            log.ip_address or "—",
            str(log.details) if log.details else "",
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 4, 60
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    write_audit(
        user=user,
        action="audit_log_exported",
        entity_type="audit",
        details={"rows": qs.count()},
    )

    filename = f"audit_log_{date.today():%Y%m%d}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
