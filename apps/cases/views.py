import logging
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db.models import DurationField, ExpressionWrapper, F, Q
from django.db.models.functions import Now
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import ListView, DetailView, FormView

from django.views.generic.edit import CreateView, UpdateView
from .forms import CaseCreateForm, CaseFilterForm, TaxpayerImportForm, TaxAuthorityDetailsForm
from .models import AdministrativeCase, Department, StagnationSettings, Taxpayer, TaxpayerType, Region, CaseBasis, CaseCategory, Position, TaxAuthorityDetails
from .services import create_case, allow_backdating
from .validators import KZValidator, IIN_BIN_ERRORS, PHONE_ERRORS

logger = logging.getLogger(__name__)


class CaseListView(LoginRequiredMixin, ListView):
    model = AdministrativeCase
    template_name = "cases/list.html"
    context_object_name = "cases"
    paginate_by = 25

    def get_queryset(self):
        qs = (
            AdministrativeCase.objects
            .for_user(self.request.user)
            .select_related("taxpayer", "responsible_user", "created_by")
            .annotate(
                days_stagnant=ExpressionWrapper(
                    Now() - F("last_activity_at"),
                    output_field=DurationField(),
                )
            )
        )
        form = CaseFilterForm(self.request.GET)
        if not form.is_valid():
            return qs

        if form.cleaned_data.get("status"):
            qs = qs.filter(status=form.cleaned_data["status"])
        if form.cleaned_data.get("department"):
            qs = qs.filter(department=form.cleaned_data["department"])
        if form.cleaned_data.get("region"):
            qs = qs.filter(region__icontains=form.cleaned_data["region"])
        if form.cleaned_data.get("responsible_user"):
            qs = qs.filter(responsible_user=form.cleaned_data["responsible_user"])
        if form.cleaned_data.get("date_from"):
            qs = qs.filter(created_at__date__gte=form.cleaned_data["date_from"])
        if form.cleaned_data.get("date_to"):
            qs = qs.filter(created_at__date__lte=form.cleaned_data["date_to"])
        if form.cleaned_data.get("search"):
            q = form.cleaned_data["search"]
            qs = qs.filter(
                Q(case_number__icontains=q)
                | Q(taxpayer__iin_bin__icontains=q)
                | Q(taxpayer__name__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter_form"] = CaseFilterForm(self.request.GET)
        context["show_dept_filter"] = self.request.user.role in ("admin", "reviewer")
        threshold = StagnationSettings.get().stagnation_days
        context["stagnation_threshold"] = threshold
        context["stagnation_warn_threshold"] = int(threshold * 0.8)
        return context


class CaseDetailView(LoginRequiredMixin, DetailView):
    model = AdministrativeCase
    template_name = "cases/detail.html"
    context_object_name = "case"

    def get_object(self, queryset=None):
        case = get_object_or_404(
            AdministrativeCase.objects.select_related("taxpayer", "responsible_user", "created_by"),
            pk=self.kwargs["pk"],
        )
        user = self.request.user
        if user.role in ("admin", "reviewer"):
            return case
        if user.role in ("operator", "observer"):
            if case.department != user.department:
                raise Http404
        if user.role == "executor":
            if case.responsible_user != user:
                raise Http404
        return case

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["events"] = (
            self.object.events
            .select_related("created_by")
            .order_by("-created_at")[:50]
        )
        # История согласований по итоговому решению дела (если есть)
        if hasattr(self.object, "final_decision"):
            from apps.approvals.services import get_history
            from apps.approvals.models import ApprovalFlow, ApprovalResult
            decision = self.object.final_decision
            context["approval_history"] = get_history(decision)
            context["pending_flow"] = (
                ApprovalFlow.objects
                .filter(entity_type="decision", entity_id=decision.pk, result=ApprovalResult.PENDING)
                .order_by("-version")
                .first()
            )
        return context


class CaseCreateView(LoginRequiredMixin, FormView):
    template_name = "cases/create.html"
    form_class = CaseCreateForm

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in ("admin", "operator"):
            messages.error(request, "У вас нет прав для создания дел.")
            return redirect("cases:list")
        return super().dispatch(request, *args, **kwargs)

    def get_initial(self):
        initial = super().get_initial()
        user = self.request.user
        if user.department_id and user.role not in ("admin", "observer"):
            initial["department"] = user.department_id
        return initial

    def form_valid(self, form):
        data = form.cleaned_data
        case = create_case(
            operator=self.request.user,
            taxpayer_data={
                "iin_bin": data["iin_bin"],
                "name": data["taxpayer_name"],
                "taxpayer_type": data["taxpayer_type"],
                "address": data.get("taxpayer_address", ""),
                "phone": data.get("taxpayer_phone", ""),
                "email": data.get("taxpayer_email", ""),
            },
            region=data["region"],
            basis=None,
            department=data.get("department"),
            category=None,
            description=data.get("description", ""),
            responsible_user=data.get("responsible_user"),
        )
        if data.get("basis"):
            case.basis.set(data["basis"])
        if data.get("category"):
            case.category.set(data["category"])
        if data.get("case_observers"):
            case.case_observers.set(data["case_observers"])
        messages.success(self.request, f"Дело {case.case_number} успешно создано.")
        return redirect("cases:detail", pk=case.pk)


class AllowBackdatingView(LoginRequiredMixin, View):
    """POST cases/<pk>/allow-backdating/ — разрешить ввод документов задним числом."""

    def post(self, request, pk):
        if request.user.role not in ("admin", "reviewer"):
            raise Http404

        case = get_object_or_404(
            AdministrativeCase.objects.for_user(request.user),
            pk=pk,
        )
        comment = request.POST.get("comment", "").strip()

        try:
            allow_backdating(case, request.user, comment)
            messages.success(
                request,
                f"Ввод документов задним числом по делу {case.case_number} разрешён."
            )
        except PermissionDenied as e:
            messages.error(request, str(e))

        return redirect("cases:detail", pk=pk)


# Маппинг русских обозначений типа НП → внутренние коды
_TAXPAYER_TYPE_MAP = {
    "юл": TaxpayerType.LEGAL,
    "юр": TaxpayerType.LEGAL,
    "фл": TaxpayerType.INDIVIDUAL,
    "физ": TaxpayerType.INDIVIDUAL,
    "ип": TaxpayerType.IE,
}


class TaxpayerImportView(LoginRequiredMixin, View):
    """GET/POST cases/taxpayers/import/ — импорт НП из Excel."""
    template_name = "cases/taxpayer_import.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in ("admin", "operator"):
            raise Http404
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        from django.shortcuts import render
        return render(request, self.template_name, {"form": TaxpayerImportForm()})

    def post(self, request):
        from django.shortcuts import render
        from apps.audit.services import audit_log

        form = TaxpayerImportForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {"form": form})

        uploaded = request.FILES["file"]
        if not uploaded.name.endswith(".xlsx"):
            messages.error(request, "Принимаются только файлы .xlsx")
            return render(request, self.template_name, {"form": form})

        try:
            import openpyxl
            wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
            ws = wb.active
        except Exception as exc:
            logger.error("TaxpayerImport: cannot open workbook: %s", exc)
            messages.error(request, f"Не удалось открыть файл: {exc}")
            return render(request, self.template_name, {"form": form})

        created_count = 0
        updated_count = 0
        errors = []
        results = []

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row_num, row in enumerate(rows, start=2):
            if not any(row):
                continue

            iin_bin = str(row[0]).strip() if row[0] is not None else ""
            name = str(row[1]).strip() if row[1] is not None else ""
            type_raw = str(row[2]).strip().lower() if row[2] is not None else ""
            address = str(row[3]).strip() if row[3] is not None else ""
            phone = str(row[4]).strip() if row[4] is not None else ""
            email = str(row[5]).strip() if row[5] is not None else ""

            # Валидация обязательных полей
            if not iin_bin or not iin_bin.isdigit() or len(iin_bin) != 12:
                errors.append({"row": row_num, "iin_bin": iin_bin, "reason": "Некорректный ИИН/БИН"})
                continue
            if not name:
                errors.append({"row": row_num, "iin_bin": iin_bin, "reason": "Пустое наименование"})
                continue

            taxpayer_type = _TAXPAYER_TYPE_MAP.get(type_raw, TaxpayerType.LEGAL)

            try:
                obj, created = Taxpayer.objects.get_or_create(
                    iin_bin=iin_bin,
                    defaults={
                        "name": name,
                        "taxpayer_type": taxpayer_type,
                        "address": address,
                        "phone": phone,
                        "email": email,
                    },
                )
                if created:
                    created_count += 1
                    results.append({"row": row_num, "iin_bin": iin_bin, "name": name, "status": "создан"})
                else:
                    # Обновляем изменяемые поля
                    changed = False
                    for field, val in [("name", name), ("address", address), ("phone", phone), ("email", email)]:
                        if val and getattr(obj, field) != val:
                            setattr(obj, field, val)
                            changed = True
                    if changed:
                        obj.save(update_fields=["name", "address", "phone", "email"])
                        updated_count += 1
                        results.append({"row": row_num, "iin_bin": iin_bin, "name": name, "status": "обновлён"})
                    else:
                        results.append({"row": row_num, "iin_bin": iin_bin, "name": name, "status": "без изменений"})
            except Exception as exc:
                logger.error("TaxpayerImport row %d error: %s", row_num, exc)
                errors.append({"row": row_num, "iin_bin": iin_bin, "reason": str(exc)})

        audit_log(
            user=request.user,
            action="taxpayer_import",
            entity_type="taxpayer",
            entity_id=0,
            details={
                "filename": uploaded.name,
                "created": created_count,
                "updated": updated_count,
                "errors": len(errors),
            },
        )

        messages.success(
            request,
            f"Импорт завершён: создано {created_count}, обновлено {updated_count}, ошибок {len(errors)}."
        )
        return render(request, self.template_name, {
            "form": TaxpayerImportForm(),
            "results": results,
            "errors": errors,
            "created_count": created_count,
            "updated_count": updated_count,
        })


def taxpayer_import_template(request):
    """GET cases/taxpayers/import/template/ — отдаёт xlsx-шаблон для заполнения."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Налогоплательщики"

    headers = ["ИИН/БИН", "Наименование / ФИО", "Тип (ЮЛ/ФЛ/ИП)", "Адрес", "Телефон", "Email"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Примеры данных
    examples = [
        ("123456789012", "ТОО «Пример»", "ЮЛ", "г. Алматы, ул. Примерная, 1", "+7 701 000 0000", "info@example.kz"),
        ("098765432109", "Иванов Иван Иванович", "ФЛ", "г. Нур-Султан, пр. Мира, 5", "+7 702 111 2233", ""),
        ("112233445566", "ИП Петров П.П.", "ИП", "", "+7 707 555 6677", "petrov@mail.kz"),
    ]
    for row_data in examples:
        ws.append(row_data)

    # Ширина колонок
    for col, width in enumerate([15, 35, 12, 35, 18, 25], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="taxpayer_import_template.xlsx"'
    return response


class ValidatePhoneView(LoginRequiredMixin, View):
    """POST cases/validate-phone/ — валидирует телефон, возвращает JSON."""

    def post(self, request):
        from django.http import JsonResponse
        phone = request.POST.get("phone", "").strip()
        result = KZValidator.validate_phone(phone)
        if result.valid:
            return JsonResponse({
                "valid": True,
                "value": result.value,
                "type": result.type,
                "metadata": result.metadata,
            })
        return JsonResponse({
            "valid": False,
            "error": PHONE_ERRORS.get(result.error, "Неверный номер телефона."),
        })


class ValidateIinView(LoginRequiredMixin, View):
    """POST cases/validate-iin/ — валидирует ИИН/БИН, возвращает JSON."""

    def post(self, request):
        from django.http import JsonResponse
        iin_bin = request.POST.get("iin_bin", "").strip()
        result = KZValidator.validate_iin_bin(iin_bin)
        if result.valid:
            return JsonResponse({
                "valid": True,
                "type": result.type,
                "metadata": result.metadata,
            })
        return JsonResponse({
            "valid": False,
            "error": IIN_BIN_ERRORS.get(result.error, "Неверный ИИН/БИН."),
        })


# ── Справочники (только admin) ────────────────────────────────────────────────

class ReferenceAdminMixin(LoginRequiredMixin):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.role != "admin":
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)


class ReferenceIndexView(ReferenceAdminMixin, View):
    def get(self, request):
        from django.shortcuts import render
        return render(request, "cases/references/list.html", {
            "counts": {
                "region": Region.objects.count(),
                "basis": CaseBasis.objects.count(),
                "category": CaseCategory.objects.count(),
                "position": Position.objects.count(),
                "department": Department.objects.count(),
                "tax_authority": TaxAuthorityDetails.objects.count(),
            }
        })


# ── Базовые generic views для справочников ────────────────────────────────────

class RefListView(ReferenceAdminMixin, ListView):
    template_name = "cases/references/ref_list.html"
    paginate_by = 50
    ref_title = ""
    create_url_name = ""   # полное имя с namespace, напр. "cases:region_create"
    update_url_name = ""
    toggle_url_name = ""
    import_url_name = ""

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(name__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["ref_title"] = self.ref_title
        ctx["create_url_name"] = self.create_url_name
        ctx["update_url_name"] = self.update_url_name
        ctx["toggle_url_name"] = self.toggle_url_name
        ctx["import_url_name"] = self.import_url_name
        ctx["q"] = self.request.GET.get("q", "")
        return ctx


class RefCreateView(ReferenceAdminMixin, CreateView):
    template_name = "cases/references/ref_form.html"
    ref_title = ""
    list_url_name = ""   # полное имя с namespace

    def get_success_url(self):
        from django.urls import reverse
        return reverse(self.list_url_name)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["ref_title"] = self.ref_title
        ctx["list_url_name"] = self.list_url_name
        ctx["action"] = "Добавить"
        return ctx


class RefUpdateView(ReferenceAdminMixin, UpdateView):
    template_name = "cases/references/ref_form.html"
    ref_title = ""
    list_url_name = ""

    def get_success_url(self):
        from django.urls import reverse
        return reverse(self.list_url_name)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["ref_title"] = self.ref_title
        ctx["list_url_name"] = self.list_url_name
        ctx["action"] = "Редактировать"
        return ctx


class RefToggleView(ReferenceAdminMixin, View):
    model = None
    list_url_name = ""

    def post(self, request, pk):
        from django.urls import reverse
        obj = get_object_or_404(self.model, pk=pk)
        obj.is_active = not obj.is_active
        obj.save(update_fields=["is_active"])
        return redirect(reverse(self.list_url_name))


class RefImportView(ReferenceAdminMixin, View):
    model = None
    list_url_name = ""
    has_legal_ref = False
    has_code = True   # Position не имеет поля code

    def post(self, request):
        import openpyxl
        from django.urls import reverse
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "Файл не выбран.")
            return redirect(reverse(self.list_url_name))

        created = updated = errors = 0
        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    if self.has_code:
                        code = str(row[0] or "").strip()
                        name = str(row[1] or "").strip() if len(row) > 1 else ""
                        legal_ref = str(row[2] or "").strip() if (self.has_legal_ref and len(row) > 2) else ""
                        if not code or not name:
                            continue
                        defaults = {"name": name}
                        if self.has_legal_ref:
                            defaults["legal_ref"] = legal_ref
                        obj, was_created = self.model.objects.get_or_create(code=code, defaults=defaults)
                    else:
                        name = str(row[0] or "").strip()
                        if not name:
                            continue
                        obj, was_created = self.model.objects.get_or_create(name=name)
                        defaults = {}

                    if was_created:
                        created += 1
                    else:
                        for k, v in defaults.items():
                            setattr(obj, k, v)
                        if defaults:
                            obj.save()
                        updated += 1
                except Exception:
                    errors += 1
            logger.info("RefImport %s: +%d ~%d err%d", self.model.__name__, created, updated, errors)
        except Exception as e:
            messages.error(request, f"Ошибка чтения файла: {e}")
            return redirect(reverse(self.list_url_name))

        messages.success(
            request,
            f"Импорт завершён: создано {created}, обновлено {updated}, ошибок {errors}."
        )
        return redirect(reverse(self.list_url_name))


# ── Region ────────────────────────────────────────────────────────────────────

class RegionListView(RefListView):
    model = Region
    template_name = "cases/references/region_list.html"
    ref_title = "Регионы"
    create_url_name = "cases:region_create"
    update_url_name = "cases:region_update"
    toggle_url_name = "cases:region_toggle"
    import_url_name = "cases:region_import"


class RegionCreateView(RefCreateView):
    model = Region
    fields = ["code", "name", "is_active"]
    ref_title = "Регионы"
    list_url_name = "cases:region_list"


class RegionUpdateView(RefUpdateView):
    model = Region
    fields = ["code", "name", "is_active"]
    ref_title = "Регионы"
    list_url_name = "cases:region_list"


class RegionToggleView(RefToggleView):
    model = Region
    list_url_name = "cases:region_list"


class RegionImportView(RefImportView):
    model = Region
    list_url_name = "cases:region_list"


# ── CaseBasis ─────────────────────────────────────────────────────────────────

class BasisListView(RefListView):
    model = CaseBasis
    template_name = "cases/references/basis_list.html"
    ref_title = "Основания дел"
    create_url_name = "cases:basis_create"
    update_url_name = "cases:basis_update"
    toggle_url_name = "cases:basis_toggle"
    import_url_name = "cases:basis_import"


class BasisCreateView(RefCreateView):
    model = CaseBasis
    fields = ["code", "name", "legal_ref", "is_active"]
    ref_title = "Основания дел"
    list_url_name = "cases:basis_list"


class BasisUpdateView(RefUpdateView):
    model = CaseBasis
    fields = ["code", "name", "legal_ref", "is_active"]
    ref_title = "Основания дел"
    list_url_name = "cases:basis_list"


class BasisToggleView(RefToggleView):
    model = CaseBasis
    list_url_name = "cases:basis_list"


class BasisImportView(RefImportView):
    model = CaseBasis
    list_url_name = "cases:basis_list"
    has_legal_ref = True


# ── CaseCategory ──────────────────────────────────────────────────────────────

class CategoryListView(RefListView):
    model = CaseCategory
    template_name = "cases/references/category_list.html"
    ref_title = "Категории дел"
    create_url_name = "cases:category_create"
    update_url_name = "cases:category_update"
    toggle_url_name = "cases:category_toggle"
    import_url_name = "cases:category_import"


class CategoryCreateView(RefCreateView):
    model = CaseCategory
    fields = ["code", "name", "is_active"]
    ref_title = "Категории дел"
    list_url_name = "cases:category_list"


class CategoryUpdateView(RefUpdateView):
    model = CaseCategory
    fields = ["code", "name", "is_active"]
    ref_title = "Категории дел"
    list_url_name = "cases:category_list"


class CategoryToggleView(RefToggleView):
    model = CaseCategory
    list_url_name = "cases:category_list"


class CategoryImportView(RefImportView):
    model = CaseCategory
    list_url_name = "cases:category_list"


# ── Position ──────────────────────────────────────────────────────────────────

class PositionListView(RefListView):
    model = Position
    template_name = "cases/references/position_list.html"
    ref_title = "Должности"
    create_url_name = "cases:position_create"
    update_url_name = "cases:position_update"
    toggle_url_name = "cases:position_toggle"
    import_url_name = "cases:position_import"


class PositionCreateView(RefCreateView):
    model = Position
    fields = ["name", "is_active"]
    ref_title = "Должности"
    list_url_name = "cases:position_list"


class PositionUpdateView(RefUpdateView):
    model = Position
    fields = ["name", "is_active"]
    ref_title = "Должности"
    list_url_name = "cases:position_list"


class PositionToggleView(RefToggleView):
    model = Position
    list_url_name = "cases:position_list"


class PositionImportView(RefImportView):
    model = Position
    list_url_name = "cases:position_list"
    has_code = False


# ── Department ────────────────────────────────────────────────────────────────

class DepartmentListView(ReferenceAdminMixin, ListView):
    model = Department
    template_name = "cases/references/department_list.html"
    paginate_by = 50

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(name__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        return ctx


class DepartmentCreateView(ReferenceAdminMixin, CreateView):
    model = Department
    fields = ["code", "name"]
    template_name = "cases/references/ref_form.html"
    success_url = reverse_lazy("cases:department_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["ref_title"] = "Подразделения"
        ctx["list_url_name"] = "cases:department_list"
        ctx["action"] = "Создать"
        return ctx


class DepartmentUpdateView(ReferenceAdminMixin, UpdateView):
    model = Department
    fields = ["code", "name"]
    template_name = "cases/references/ref_form.html"
    success_url = reverse_lazy("cases:department_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["ref_title"] = "Подразделения"
        ctx["list_url_name"] = "cases:department_list"
        ctx["action"] = "Редактировать"
        return ctx


# ── Реквизиты КГД ─────────────────────────────────────────────────────────────

class TaxAuthorityListView(ReferenceAdminMixin, ListView):
    model = TaxAuthorityDetails
    template_name = "cases/references/tax_authority_list.html"
    paginate_by = 50

    def get_queryset(self):
        return TaxAuthorityDetails.objects.select_related("department").order_by("name")


class TaxAuthorityCreateView(ReferenceAdminMixin, CreateView):
    model = TaxAuthorityDetails
    form_class = TaxAuthorityDetailsForm
    template_name = "cases/references/tax_authority_form.html"
    success_url = reverse_lazy("cases:tax_authority_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["action"] = "Создать"
        return ctx

    def form_valid(self, form):
        from apps.audit.services import audit_log
        obj = form.save(commit=False)
        obj.updated_by = self.request.user
        obj.save()
        audit_log(
            user=self.request.user,
            action="tax_authority_created",
            entity_type="tax_authority_details",
            entity_id=obj.pk,
            details={"name": obj.name},
        )
        return redirect(self.success_url)


class TaxAuthorityUpdateView(ReferenceAdminMixin, UpdateView):
    model = TaxAuthorityDetails
    form_class = TaxAuthorityDetailsForm
    template_name = "cases/references/tax_authority_form.html"
    success_url = reverse_lazy("cases:tax_authority_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["action"] = "Редактировать"
        return ctx

    def form_valid(self, form):
        from apps.audit.services import audit_log
        obj = form.save(commit=False)
        obj.updated_by = self.request.user
        obj.save()
        audit_log(
            user=self.request.user,
            action="tax_authority_updated",
            entity_type="tax_authority_details",
            entity_id=obj.pk,
            details={"name": obj.name},
        )
        return redirect(self.success_url)
