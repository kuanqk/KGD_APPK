"""
Экспорт отчётов в PDF (WeasyPrint) и XLSX (openpyxl).
Каждый экспортёр логирует действие в AuditLog.
"""
import io
import logging
from datetime import date

from django.http import HttpResponse
from django.utils import timezone

from apps.audit.services import audit_log

logger = logging.getLogger(__name__)

# Метаданные отчётов: название, столбцы для XLSX, поля из QS
REPORT_META = {
    "cases_by_status": {
        "title": "Дела по статусам",
        "type": "dict",
    },
    "cases_by_region": {
        "title": "Дела по регионам",
        "type": "dict",
    },
    "overdue_cases": {
        "title": "Просроченные дела",
        "type": "cases_qs",
        "columns": ["Номер дела", "Налогоплательщик", "Регион", "Ответственный", "Статус", "Дата создания"],
    },
    "terminated_cases": {
        "title": "Прекращённые дела",
        "type": "cases_qs",
        "columns": ["Номер дела", "Налогоплательщик", "Регион", "Ответственный", "Дата закрытия"],
    },
    "audit_initiated_cases": {
        "title": "Дела с назначенной проверкой",
        "type": "cases_qs",
        "columns": ["Номер дела", "Налогоплательщик", "Регион", "Ответственный", "Дата закрытия"],
    },
    "avg_case_duration": {
        "title": "Среднее время жизни дел",
        "type": "dict",
    },
    "revision_journal": {
        "title": "Журнал возвратов на доработку",
        "type": "approvals_qs",
        "columns": ["Тип", "ID", "Итерация", "Направил", "Дата направления", "Рецензент", "Дата рассмотрения", "Комментарий"],
    },
    "discipline_report": {
        "title": "Дисциплинарный отчёт",
        "type": "discipline",
        "columns": ["Сотрудник", "Назначено", "Просрочено", "Завершено"],
    },
    "cases_registry": {
        "title": "Реестр дел",
        "type": "cases_qs",
        "columns": ["Номер дела", "Налогоплательщик", "ИИН/БИН", "Регион", "Подразделение", "Ответственный", "Статус", "Дата создания"],
    },
}


def _get_report_data(report_type: str, filters: dict, user):
    """Вызывает нужную функцию из services и возвращает данные."""
    from apps.reports import services as svc
    fn_map = {
        "cases_by_status": svc.cases_by_status,
        "cases_by_region": svc.cases_by_region,
        "overdue_cases": svc.overdue_cases,
        "terminated_cases": svc.terminated_cases,
        "audit_initiated_cases": svc.audit_initiated_cases,
        "avg_case_duration": svc.avg_case_duration,
        "revision_journal": svc.revision_journal,
        "discipline_report": svc.discipline_report,
        "cases_registry": svc.cases_registry,
    }
    fn = fn_map.get(report_type)
    if fn is None:
        raise ValueError(f"Неизвестный отчёт: {report_type}")
    return fn(filters, user)


def _rows_from_data(report_type: str, data) -> list:
    """Преобразует данные отчёта в список строк для таблицы."""
    meta = REPORT_META[report_type]
    kind = meta["type"]

    if kind == "dict":
        return [[k, v] for k, v in data.items()]

    if kind == "cases_qs":
        rows = []
        for case in data:
            if report_type == "cases_registry":
                rows.append([
                    case.case_number,
                    case.taxpayer.name,
                    case.taxpayer.iin_bin,
                    case.region,
                    case.department or "—",
                    str(case.responsible_user or "—"),
                    case.get_status_display(),
                    case.created_at.strftime("%d.%m.%Y") if case.created_at else "—",
                ])
            elif report_type in ("terminated_cases", "audit_initiated_cases"):
                rows.append([
                    case.case_number,
                    case.taxpayer.name,
                    case.region,
                    str(case.responsible_user or "—"),
                    case.closed_at.strftime("%d.%m.%Y") if case.closed_at else "—",
                ])
            else:
                rows.append([
                    case.case_number,
                    case.taxpayer.name,
                    case.region,
                    str(case.responsible_user or "—"),
                    case.get_status_display(),
                    case.created_at.strftime("%d.%m.%Y") if case.created_at else "—",
                ])
        return rows

    if kind == "approvals_qs":
        rows = []
        for flow in data:
            rows.append([
                flow.get_entity_type_display(),
                flow.entity_id,
                flow.version,
                str(flow.sent_by or "—"),
                flow.sent_at.strftime("%d.%m.%Y %H:%M") if flow.sent_at else "—",
                str(flow.reviewed_by or "—"),
                flow.reviewed_at.strftime("%d.%m.%Y %H:%M") if flow.reviewed_at else "—",
                flow.comment,
            ])
        return rows

    if kind == "discipline":
        return [
            [
                str(row["user"]),
                row["assigned"],
                row["overdue"],
                row["completed"],
            ]
            for row in data
        ]

    return []


def export_xlsx(report_type: str, filters: dict, user) -> HttpResponse:
    """Генерирует XLSX и возвращает HttpResponse."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl not installed — cannot export XLSX")
        raise ValueError("Для экспорта XLSX необходим пакет openpyxl.")

    meta = REPORT_META.get(report_type)
    if meta is None:
        raise ValueError(f"Неизвестный отчёт: {report_type}")

    data = _get_report_data(report_type, filters, user)
    rows = _rows_from_data(report_type, data)
    columns = meta.get("columns") or ["Ключ", "Значение"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = meta["title"][:31]

    # Заголовок отчёта
    ws.merge_cells(f"A1:{chr(64 + len(columns))}1")
    title_cell = ws["A1"]
    title_cell.value = meta["title"]
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    ws.append([f"Сформировано: {date.today():%d.%m.%Y}  |  Пользователь: {user}"])
    ws.append([])

    # Заголовки столбцов
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(columns)
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Данные
    for row in rows:
        ws.append(row)

    # Авто-ширина колонок
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    audit_log(
        user=user,
        action="report_exported_xlsx",
        entity_type="report",
        details={"report_type": report_type, "filters": str(filters)},
    )

    filename = f"{report_type}_{date.today():%Y%m%d}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_pdf(report_type: str, filters: dict, user) -> HttpResponse:
    """Генерирует PDF через WeasyPrint и возвращает HttpResponse."""
    from django.template.loader import render_to_string

    meta = REPORT_META.get(report_type)
    if meta is None:
        raise ValueError(f"Неизвестный отчёт: {report_type}")

    data = _get_report_data(report_type, filters, user)
    rows = _rows_from_data(report_type, data)
    columns = meta.get("columns") or ["Ключ", "Значение"]

    html = render_to_string("reports/pdf_template.html", {
        "title": meta["title"],
        "columns": columns,
        "rows": rows,
        "generated_at": timezone.now(),
        "user": user,
        "filters": filters,
    })

    try:
        from io import BytesIO
        from xhtml2pdf import pisa
        buffer = BytesIO()
        pisa.CreatePDF(html.encode("utf-8"), dest=buffer)
        pdf_bytes = buffer.getvalue()
    except Exception as exc:
        logger.error("xhtml2pdf ошибка: %s — возвращаем HTML", exc)
        pdf_bytes = html.encode("utf-8")
        audit_log(
            user=user,
            action="report_exported_pdf",
            entity_type="report",
            details={"report_type": report_type, "fallback": True},
        )
        filename = f"{report_type}_{date.today():%Y%m%d}.html"
        response = HttpResponse(pdf_bytes, content_type="text/html; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    audit_log(
        user=user,
        action="report_exported_pdf",
        entity_type="report",
        details={"report_type": report_type, "filters": str(filters)},
    )

    filename = f"{report_type}_{date.today():%Y%m%d}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
